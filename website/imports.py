from flask import Blueprint,request,render_template,redirect,url_for
from openpyxl import load_workbook
from website import db_connect
from .views import get_data
from datetime import datetime
import re

imports = Blueprint('imports',__name__)

def get_partial_amount(percent,total):
    return round((float(percent)/100)*float(total),2)

def vehicle_plate_check(plate):
    conn = db_connect()
    cur = conn.cursor()
    if len(plate) not in (9,10):
        return f"Invalid Length of vehicle plate at Row "
    else:
        state_code = plate[:3]
        cur.execute("SELECT id,short_name FROM state;")
        state_codes = {data[1]:data[0] for data in cur.fetchall()}
        if state_code not in state_codes:
            return f"Invald State Code at Row "
        elif 'UN' in plate:
            return 1
        elif not plate[-4:].isdigit():
            return f"Invalid Last Digits Plate at Row "
        return 1


@imports.route("/excel",methods=['GET','POST'])
def excel_import():
    if not request.cookies.get('user_roles') or not request.cookies.get('pg-username'):
        return redirect(url_for('views.home'))
    if request.method == 'POST':
        upload_file = request.files['upload_serivce_datas']
        excel_file_type = request.form.get('selectExcelFile')
        view_type = request.form.get('selectView')

        if upload_file.filename != '' and upload_file.filename.endswith(".xlsx"):
            workbook = load_workbook(filename=upload_file,data_only=True,read_only=True)
            try:
                worksheet = workbook[excel_file_type]
            except:
                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f'The sheet name of excel import file must be <strong>{excel_file_type}..<&#47;strong>'))
            conn = db_connect()
            cur = conn.cursor()
            if excel_file_type == 'Jobs Data':
                eachJob_insert_query = """INSERT INTO eachJob (job_date,job_no,business_unit_id,shop_id,invoice_no,customer_id,vehicle_id,job_name,
                job_type_id,total_amt,fst_pic_id,fst_pic_amt,sec_pic_id,sec_pic_amt,thrd_pic_id,thrd_pic_amt,frth_pic_id,
                frth_pic_amt,lst_pic_id,lst_pic_amt,eachJob_concatenated,pic_rate_id,ownership_id) VALUES """
                cur.execute("SELECT id,name,shop_id FROM technicians;")
                technicians = cur.fetchall()
                technicians = {(data[1].lower(),data[2]):data[0] for data in technicians}
                cur.execute("SELECT id,name FROM jobType;")
                job_types = cur.fetchall()
                job_types = {data[1].lower():data[0] for data in job_types}
                cur.execute("SELECT fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate,id FROM pic ORDER BY id;")
                all_rates = {data[5]:[data[:5],data[6]] for data in cur.fetchall()}
                cur.execute("SELECT shop.business_unit_id,shop.id as shop_id,LOWER(unit.name),LOWER(shop.name) FROM res_partner AS unit INNER JOIN shop ON shop.business_unit_id = unit.id;")
                unit_shop_datas = cur.fetchall()
                unit_shop_dct = {data[2:]:data[:2] for data in unit_shop_datas}
                conflict_unique_column = "eachJob_concatenated"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    # check invalid rate
                    if row[19].value is None or row[19].value.strip() == "" or row[19].value.strip() not in all_rates:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Unregistered PIC Rate at row {row_counter}"))
                    # check all invalid fields
                    if None in (row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,
                                row[8].value,row[9].value,row[10].value,row[11].value,row[17].value,row[18].value,row[19].value) or "" in (row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,
                                row[8].value,row[9].value,row[10].value,row[11].value,row[17].value,row[18].value,row[19].value):
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))
                    # get unit and shop id
                    unit_shop_ids = unit_shop_dct.get((row[17].value.strip().lower(),row[18].value.strip().lower()))
                    if not unit_shop_ids:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank Business Unit at row {row_counter}"))
                    elif str(unit_shop_ids[1]) not in request.cookies.get("user_roles").split(","):
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"You don't have enough access to import for {row[18].value.strip()}"))  
                    # get rate and assign tech name 
                    rate = all_rates[row[19].value.strip()]
                    cell_counter = 12
                    tech_names = []
                    for each_rate in rate[0]:
                        if each_rate:
                            if row[cell_counter].value is None or row[cell_counter].value.strip() == "":
                                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Blank technician name at row {row_counter} and column {cell_counter+1}, must be matched with pic_rate..."))
                            if (row[cell_counter].value.strip().lower(),unit_shop_ids[1]) not in technicians:
                                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Technician Name at row - {row_counter} and column {cell_counter+1}.. \nPlease check or add names in Configurations -> Technicinans.."))
                            tech_names.append((row[cell_counter].value.strip().lower(),unit_shop_ids[1]))
                        else:
                            tech_names.append(('bot',None))
                        cell_counter += 1 
                    # get brand / model id
                    cur.execute("SELECT id,brand_id FROM vehicle_model WHERE LOWER(name) = %s;",(row[6].value.strip().lower(),))
                    idds = cur.fetchone()
                    if not idds:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Unregistered Vehicle Model at row {row_counter}"))  
                    # check vehicle plate
                    returned_mgs = vehicle_plate_check(row[5].value.strip())
                    if returned_mgs != 1:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"{returned_mgs} {row_counter}."))
                    # create vehicle
                    vehicle_year = row[7].value
                    # check vehicle
                    plate = row[5].value.strip() if 'UN' in row[5].value.strip() else row[5].value.strip().replace("-","")
                    cur.execute("SELECT id FROM vehicle WHERE REPLACE(LOWER(plate),'-','') = %s;",(plate,))
                    vehicle_datas = cur.fetchone()                    
                    phone = re.sub(r'\D', '', str(row[4].value).strip().split(",")[0])
                    if not phone.startswith("0"):
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Phone {phone} at {row_counter}.\n Please Starts  customer phone with 09.."))   
                    cur.execute("SELECT id,name,phone,secondary_phone FROM customer WHERE phone = %s or secondary_phone = %s or name = %s;",(phone,phone,row[3].value.strip().upper()))
                    cus_datas = cur.fetchall()
                    if cus_datas:
                        for cus_data in cus_datas:
                            if cus_data[2] == phone or cus_data[3] == phone:
                                cus_id = cus_data[0]
                            if vehicle_datas:
                                cur.execute("SELECT customer_id FROM ownership WHERE vehicle_id = %s;",(vehicle_datas))
                                ownership_datas_for_vehicle_data = cur.fetchall()
                                for ownership_data in ownership_datas_for_vehicle_data:
                                    if ownership_data[0] == cus_data[0]:
                                        cur.execute("UPDATE customer SET secondary_phone = %s WHERE id = %s;",(phone,ownership_data[0]))
                                        cus_id = ownership_data[0]
                                        break
                            if cus_id:
                                break
                        if not cus_id:                       
                            cur.execute(""" INSERT INTO customer (name,phone)
                                    VALUES (%s,%s)
                                    ON CONFLICT (phone) DO NOTHING 
                                    RETURNING id """,(row[3].value.strip().upper(),phone))
                            cus_id = cur.fetchone()[0] 
                    else:
                        # create new customer
                        cur.execute(""" INSERT INTO customer (name,phone)
                                        VALUES (%s,%s)
                                        ON CONFLICT (phone) DO NOTHING 
                                        RETURNING id """,(row[3].value.strip().upper(),phone))
                        cus_id = cur.fetchone()[0]
                    if vehicle_datas:
                        vehicle_id = vehicle_datas[0]
                    else:
                        cur.execute(""" INSERT INTO vehicle (UPPER(plate),model_id,brand_id,year)
                                        VALUES (%s,%s,%s,%s)
                                        ON CONFLICT (plate) DO UPDATE
                                        SET plate = EXCLUDED.plate
                                        RETURNING id;""",(plate.upper(),idds[0],idds[1],vehicle_year))
                        vehicle_id = cur.fetchone()[0]
                    if not cus_id:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Data for Customer ({row[3].value.strip()}-{phone}) and Vehicle ({row[5].value.strip()}) at -  {row_counter} "))            
                    # create ownership
                    cur.execute(""" WITH inserted AS (
                                        INSERT INTO ownership (vehicle_id,customer_id,unique_owner)
                                        VALUES (%s,%s,%s)
                                        ON CONFLICT (unique_owner) DO NOTHING 
                                        RETURNING id
                                    )
                                    SELECT id FROM inserted
                                    UNION ALL
                                    SELECT id FROM ownership WHERE unique_owner = %s 
                                    LIMIT 1;""",(vehicle_id,cus_id,f"{vehicle_id}-{cus_id}",f"{vehicle_id}-{cus_id}"))
                    ownership_id = cur.fetchall()[0][0]
                    # create psfu 
                    cur.execute("INSERT INTO psfu (job_no,job_date,shop_id,psfu_concatenated) VALUES (%s,%s,%s,%s) ON CONFLICT (psfu_concatenated) DO NOTHING;",(row[2].value,row[1].value.strftime("%Y/%m/%d"),str(unit_shop_ids[1]),row[2].value+row[1].value.strftime("%Y/%m/%d")+str(unit_shop_ids[1])))
                    # check job type
                    if row[10].value.strip().lower() not in job_types:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Job Type at row -  {row_counter}"))                    
                    eachJob_concatenated = row[1].value.strftime("%Y/%m/%d") + row[2].value + row[9].value + str(unit_shop_ids[1])
                    # extends query
                    try:
                        eachJob_insert_query += f"""('{row[1].value}','{row[2].value}','{unit_shop_ids[0]}','{unit_shop_ids[1]}','{row[8].value}','{cus_id}','{vehicle_id}','{row[9].value}','{job_types[row[10].value.strip().lower()]}','{int(row[11].value):.2f}','{technicians[tech_names[0]]}','{get_partial_amount(rate[0][0],row[11].value)}','{technicians[tech_names[1]]}','{get_partial_amount(rate[0][1],row[11].value)}','{technicians[tech_names[2]]}','{get_partial_amount(rate[0][2],row[11].value)}','{technicians[tech_names[3]]}','{get_partial_amount(rate[0][3],row[11].value)}','{technicians[tech_names[4]]}','{get_partial_amount(rate[0][4],row[11].value)}','{eachJob_concatenated}','{rate[1]}','{ownership_id}'),"""
                    except ValueError:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Job Data at row -  {row_counter}"))
                cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
                # conn.commit()    
                cur.close()
                conn.close()
                return redirect(url_for('views.show_service_datas',typ='service-datas'))            
            elif excel_file_type == 'Types Data':
                eachJob_insert_query = "INSERT INTO jobType (name) VALUES "
                conflict_unique_column = "name"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    eachJob_insert_query += f"('{row[0].value}'),"
            elif excel_file_type == 'Technicians': 
                eachJob_insert_query = "INSERT INTO technicians (name,business_unit_id,shop_id) VALUES "
                conflict_unique_column = "name"
                cur.execute("SELECT name,business_unit_id,id FROM shop;")
                shops_dct = {data[0]:data[1:] for data in cur.fetchall()}
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    shop_name = row[2].value.strip()
                    if None in (row[0].value,row[1].value,row[2].value) or row[0].value.strip() == "" or row[1].value.strip() == "" or shop_name == "" or shop_name not in shops_dct:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))                           
                    eachJob_insert_query += f"('{row[0].value.strip().upper()}','{shops_dct[shop_name][0]}','{shops_dct[shop_name][1]}'),"
            elif excel_file_type == 'brand-model':
                eachJob_insert_query = "INSERT INTO vehicle_model(brand_id,name) VALUES"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    if None in (row[0].value,row[1].value) or row[0].value.strip() == "" or row[1].value.strip() == "":
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))        
                    cur.execute(f"INSERT INTO vehicle_brand (name) VALUES('{row[0].value.strip()}') ON CONFLICT (name) DO UPDATE set name = '{row[0].value.strip()}' RETURNING id;")
                    eachJob_insert_query += f"({cur.fetchall()[0][0]},'{row[1].value}'),"
                conflict_unique_column = "name"
            cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
            conn.commit()
            cur.close()
            conn.close()
        else:
            return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f'<strong> Invalid Excel File Type.. <&#47;strong>'))
    return redirect(url_for('views.show_service_datas',typ=view_type))

@imports.route("/create-form/<typ>")
def show_create_form(typ,mgs=None):
    if not request.cookies.get('user_roles') or not request.cookies.get('pg-username'):
        return redirect(url_for('views.home'))
    mgs = request.args.get("mgs")
    conn = db_connect()
    cur = conn.cursor()
    result = ["","",""]
    if typ == 'service-datas':
        user_roles = tuple(request.cookies.get("user_roles").split(","))
        cur.execute("SELECT id,name FROM technicians WHERE shop_id in %s;",(user_roles,))
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM jobType;")
        result.append(cur.fetchall())
        cur.execute("SELECT unique_rate FROM pic;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,short_name,name FROM state;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM vehicle_brand;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
        cur.execute("SELECT name FROM customer;")
        result.append(cur.fetchall())
    if typ == 'check-in-out':
        user_roles = tuple(request.cookies.get("user_roles").split(","))
        cur.execute("SELECT id,name FROM technicians WHERE shop_id in %s;",(user_roles,))
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM jobType;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,short_name,name FROM state;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM vehicle_brand;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
    elif typ == 'customers-create':
        result = [datetime.now().strftime("%Y-%m-%d")]
        cur.execute("SELECT 'CUS'||LPAD((id+1)::text,7,'0') FROM customer ORDER BY id DESC LIMIT 1;")
        try:
            result.append(cur.fetchall()[0][0])
        except IndexError:
            result.append('CUS0000001')
        cur.execute("SELECT id,name FROM state;")
        result.append(cur.fetchall())
        return render_template('registration_form.html',result=result,typ=typ)
    elif typ == 'vehicles-create':
        result = [datetime.now().strftime("%Y-%m-%d")]
        cur.execute("SELECT name FROM vehicle_brand;")
        result.append(cur.fetchall())
        cur.execute("SELECT short_name FROM state;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
        return render_template('registration_form.html',result=result,typ=typ)  

    return render_template('input_form.html',result=result,mgs=mgs,typ=typ)

@imports.route("/keep-in-import/<typ>",methods=['POST'])
def keep_in_import(typ): 
    if not request.cookies.get('user_roles') or not request.cookies.get('pg-username'):
        return redirect(url_for('views.home'))       
    mgs = None
    if request.method == 'POST':
        conn = db_connect()
        cur = conn.cursor()
        if typ == 'service-datas':
            conflict_unique_column = 'eachjob_concatenated'
            eachJob_insert_query = """ INSERT INTO eachJob (job_date,job_no,business_unit_id,shop_id,invoice_no,customer_id,vehicle_id,job_name,
                job_type_id,total_amt,fst_pic_id,fst_pic_amt,sec_pic_id,sec_pic_amt,thrd_pic_id,thrd_pic_amt,frth_pic_id,
                frth_pic_amt,lst_pic_id,lst_pic_amt,eachJob_concatenated,pic_rate_id,ownership_id) VALUES """
            edit = request.form.get("newOrEdit")
            #
            job_no = request.form.get("jobNo")
            invoice_no = request.form.get("invoiceNo")
            shop_id = request.form.get("shop")
            cur.execute("SELECT id FROM eachJob WHERE (job_no = %s and shop_id = %s) or ( shop_id = %s and invoice_no = %s);",(job_no,shop_id,shop_id,invoice_no))
            if len(cur.fetchall()) != 0 and not edit:
                mgs = 'Job No. / Invoice No. is already existed in our system...'
            else:
                job_date = request.form.get("jobDate")
                cur.execute("SELECT business_unit_id FROM shop WHERE id = %s;",(shop_id,))
                unit_id = cur.fetchall()[0][0]
                #
                descriptions = request.form.getlist("description")[1:]
                job_types = [data.strip() for data in request.form.getlist('jobType')[1:]]
                cur.execute("SELECT fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate,id FROM pic;")
                all_rates = {data[5]:[data[:5],data[6]] for data in cur.fetchall()}
                cur.execute("SELECT id,name FROM technicians;")
                all_technicians = {data[1]:data[0] for data in cur.fetchall()}
                #
                pic_ones = [data for data in request.form.getlist("picOne")[1:] ]
                pic_twos = [data for data in request.form.getlist("picTwo")[1:] ]
                pic_threes = [data for data in request.form.getlist("picThree")[1:] ]
                pic_fours = [data for data in request.form.getlist("picFour")[1:] ]
                pic_fives = [data for data in request.form.getlist("picFive")[1:] ]
                pic_rates = [data.split(",") for data in request.form.getlist("pic-rate")]
                job_costs = request.form.getlist("jobCost")[1:]
                if edit:
                    old_job_no = request.form.get("oldJobNo")
                    print(get_data('eachJobDelForm',old_job_no))
                vehicle_id = request.form.get("vehicleInformation")
                new_onwership = False
                if request.form.get("vehicleInformation") == "None":
                    plate = request.form.get("regState") + request.form.get("regPrefix") + request.form.get("regDigits")
                    model = request.form.get("model_id")
                    brand = request.form.get("brand_id")
                    year = request.form.get("year")
                    cur.execute("INSERT INTO vehicle (plate,model_id,brand_id,year) VALUES (%s,%s,%s,%s) RETURNING id;",(plate,model,brand,year)) 
                    vehicle_id = cur.fetchall()[0][0]     
                    new_onwership = True    
                customer_id = request.form.get("customerInformation")
                if request.form.get("customerInformation") == "None":
                    cus_name = request.form.get("customerName")
                    address = request.form.get("fullAddress")
                    state_id = request.form.get("state")
                    phone = request.form.get("phone")
                    cur.execute("INSERT INTO customer (name,address,phone,state_id) VALUES (%s,%s,%s,%s) ON CONFLICT (phone) DO NOTHING RETURNING id;",(cus_name,address,phone,state_id))
                    customer_datas = cur.fetchall()
                    if customer_datas:
                        customer_id = customer_datas[0][0]
                        new_onwership = True
                    else:
                        return redirect(url_for('imports.show_create_form',typ='service-datas',mgs='Customer with same phone number is existed..'))
                if new_onwership:
                    cur.execute("INSERT INTO ownership (customer_id,vehicle_id,start_date,unique_owner) VALUES (%s,%s,%s,%s) RETURNING id;",(customer_id,vehicle_id,datetime.now().strftime("%Y-%m-%d"),str(vehicle_id)+'-'+str(customer_id)))
                else:
                    cur.execute("SELECT id FROM ownership WHERE customer_id = %s and vehicle_id = %s;",(customer_id,vehicle_id))                
                ownership_id = cur.fetchall()[0][0]
                cur.execute("INSERT INTO psfu (job_no,job_date,shop_id,psfu_concatenated) VALUES (%s,%s,%s,%s) ON CONFLICT (psfu_concatenated) DO NOTHING;",(job_no,job_date,shop_id,job_no+job_date+shop_id))
                for data in zip(descriptions,job_types,pic_ones,pic_twos,pic_threes,pic_fours,pic_fives,job_costs,pic_rates):
                    eachJob_concatenated = job_date.replace("-","/") + job_no + data[0]
                    eachJob_insert_query += f"""('{job_date}','{job_no}','{unit_id}','{shop_id}','{invoice_no}','{customer_id}','{vehicle_id}','{data[0]}',
                    '{data[1]}','{data[7]}','{all_technicians.get(data[2],0)}','{get_partial_amount(data[8][0],data[7])}','{all_technicians.get(data[3],0)}','{get_partial_amount(data[8][1],data[7])}','{all_technicians.get(data[4],0)}','{get_partial_amount(data[8][2],data[7])}','{all_technicians.get(data[5],0)}','{get_partial_amount(data[8][3],data[7])}','{all_technicians.get(data[6],0)}','{get_partial_amount(data[8][4],data[7])}','{eachJob_concatenated}','{all_rates[",".join(data[8])][1]}','{ownership_id}'),"""
                cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
                conn.commit()
            return redirect(url_for('imports.show_create_form',typ='service-datas',mgs=mgs))
        elif typ == 'check-in-out':
            vehicle_id = request.form.get("vehicleInformation")
            job_no = request.form.get("jobNo")
            job_date = request.form.get("jobDate")
            shop_id = request.form.get("shop")
            cur.execute("SELECT id FROM eachJob WHERE (job_no = %s and job_date = %s);",(job_no,job_date))
            if len(cur.fetchall()) != 0 and not edit:
                mgs = 'Job No. / Invoice No. is already existed in our system...'
            else:
                cur.execute("SELECT business_unit_id FROM shop WHERE id = %s;",(shop_id,))
                unit_id = cur.fetchall()[0][0]
                #
                descriptions = request.form.getlist("description")[1:]
            print(vehicle_id)
            return ""
        elif typ == 'pic-rate':
            idd = request.form.get("idd")
            rates = request.form.getlist('rate')
            if idd:
                rates = rates[5:]
            rates = [rate.strip() for rate in rates]
            rates.append(','.join(rates))
            if not idd:
                cur.execute(f"INSERT INTO pic (fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate) VALUES {tuple(rates)} ON CONFLICT (unique_rate) DO NOTHING;")
            else:
                cur.execute(f""" UPDATE pic SET fst_rate = {rates[0]},sec_rate = {rates[1]},thrd_rate = {rates[2]},frth_rate = {rates[3]},lst_rate = {rates[4]},unique_rate = '{rates[5]}' WHERE id = {idd} AND NOT EXISTS (SELECT 1 FROM pic WHERE unique_rate = '{rates[5]}'); """)
        elif typ == 'technician':
            idd = request.form.get("idd")
            tech_name  = request.form.getlist("tech")
            shop_name = request.form.getlist("shop")
            shop_name = shop_name[1] if idd else shop_name[0] 
            cur.execute("SELECT business_unit_id,id FROM shop WHERE name = %s;",(shop_name,))
            data = cur.fetchall()
            if idd:
                if len(data) == 0:
                    mgs = 'Invalid Shop / Business Unit'
                else:
                    cur.execute(f""" UPDATE technicians SET name = '{tech_name[1]}',business_unit_id = '{data[0][0]}',shop_id = '{data[0][1]}' WHERE id = '{idd}' and NOT EXISTS ( SELECT 1 FROM technicians WHERE name = '{tech_name[1]}' and business_unit_id = '{data[0][0]}' and shop_id = '{data[0][1]}') ;""")
            else:
                if len(data) == 0:
                    mgs = 'Invalid Shop / Business Unit'
                else:
                    cur.execute(f""" INSERT INTO technicians (name,business_unit_id,shop_id) VALUES ('{tech_name[0]}','{data[0][0]}','{data[0][1]}') ON CONFLICT (name) DO NOTHING;""")
        elif typ == 'jobType':
            idd = request.form.get("idd")
            job_type = request.form.getlist("jobType")
            if idd:
                cur.execute(f""" UPDATE jobType SET name = '{job_type[1]}' WHERE id = '{idd}' and NOT EXISTS ( SELECT 1 FROM jobtype WHERE name = '{job_type[1]}');""")
            else:
                 cur.execute(f""" INSERT INTO jobType (name) VALUES ('{job_type[0]}') ON CONFLICT (name) DO NOTHING;""")
        elif typ == 'customers':
            reg_date = request.form.get("register-date")
            name = request.form.get("name")
            state = request.form.get("state")
            address = request.form.get("address")
            phone = request.form.get("phone")
            vehicle_ids = request.form.getlist("vehicleIds")
            customer_id = request.form.get("customerId")
            unique_phone_query = f"AND id <> '{customer_id}'" if customer_id else ""
            cur.execute(f"SELECT id FROM customer WHERE phone = '{phone}' {unique_phone_query};")     
            if customer_id:
                if len(cur.fetchall()) == 0:
                    cur.execute("UPDATE customer SET name = %s,state_id = %s,address = %s,phone = %s,registered_date = %s WHERE id = %s;",(name,state,address,phone,reg_date,customer_id))
                else:
                    mgs = f"Phone - {phone} is already existed.."                    
                for vehicle_id in vehicle_ids:
                    cur.execute("INSERT INTO customer (vehicle_id,customer_id,unique_owner) VALUES (%s,%s,%s) ON CONFLICT (unique_owner) DO NOTHING;",(vehicle_id,customer_id,vehicle_id + '-' +customer_id))     
            else:
                if len(cur.fetchall()) == 0:
                    cur.execute("INSERT INTO customer (name,address,phone,state_id,registered_date) VALUES (%s,%s,%s,%s,%s);",(name,address,phone,state,reg_date))
                else:
                    mgs = f"Phone - {phone} is already existed.."
        elif typ == 'vehicles':
            vehicle_id = request.form.get("vehicle-id")
            reg_date = request.form.get("register-date")
            plate = request.form.get("fst-part") + request.form.get("sec-part") + request.form.get("thrd-part")
            brand_id = request.form.get("brand_id")
            model_id = request.form.get("model_id")
            year = request.form.get("year")
            unique_plate_query = f"AND id <> '{vehicle_id}'" if vehicle_id else ""
            cur.execute(f"SELECT id FROM vehicle WHERE plate = '{plate}' {unique_plate_query};")  
            if vehicle_id:
                if len(cur.fetchall()) == 0:
                    cur.execute("UPDATE vehicle SET register_date = %s,plate = %s,brand_id = %s,model_id = %s,year = %s WHERE id = %s;",(reg_date,plate,brand_id,model_id,year,vehicle_id))
                else:
                    mgs = f"Plate - {plate} is already registered in our system.."
            else:
                if len(cur.fetchall()) == 0:
                    cur.execute("INSERT INTO vehicle (plate,brand_id,model_id,year,register_date) VALUES (%s,%s,%s,%s,%s);",(plate,brand_id,model_id,year,reg_date))
                else:
                    mgs = f"Plate - {plate} is already registered in our system.."
        elif typ == 'brand':
            brand_name = request.form.get("brand").strip()
            model_name = request.form.get("model").strip()
            cur.execute("INSERT INTO vehicle_brand(name) VALUES(%s) ON CONFLICT (name) DO UPDATE SET name = %s RETURNING id;",(brand_name,brand_name))
            brand_id = cur.fetchall()[0][0]
            cur.execute("SELECT id FROM vehicle_model WHERE brand_id = %s AND name = %s;",(brand_id,model_name))
            if len(cur.fetchall()) == 0:
                cur.execute("INSERT INTO vehicle_model(brand_id,name) VALUES (%s,%s);",(brand_id,model_name))
        conn.commit()
    return redirect(url_for('views.show_service_datas',typ=typ,mgs=mgs))
